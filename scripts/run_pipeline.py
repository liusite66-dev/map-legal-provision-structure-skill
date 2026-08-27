#!/usr/bin/env python3
"""Create auditable hierarchy and citation graphs from supplied PRC legal texts."""
from __future__ import annotations

import argparse
import hashlib
import html
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

SUPPORTED = {".docx", ".doc", ".pdf", ".txt", ".md", ".html", ".htm", ".xlsx", ".xls", ".pptx", ".ppt", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}
ARCHIVES = {".zip"}
FORBIDDEN = {".exe", ".dll", ".dylib", ".sh", ".bat", ".cmd", ".js", ".jar", ".py"}
MARKER = ".legal-structure-workspace"
SCHEMA = "legal-provision-structure/v1"
MAX_FILES, MAX_FILE, MAX_TOTAL, MAX_RATIO = 2000, 100 * 1024 * 1024, 1024 * 1024 * 1024, 100
NUM = "一二三四五六七八九十百千万零〇两0-9"
HEADING = [("编", re.compile(rf"^第([{NUM}]+)编(?:\s*(.*))?$")), ("章", re.compile(rf"^第([{NUM}]+)章(?:\s*(.*))?$")), ("节", re.compile(rf"^第([{NUM}]+)节(?:\s*(.*))?$")), ("条", re.compile(rf"^第([{NUM}]+)条(?:\s*(.*))?$"))]
EXPLICIT_CLAUSE = re.compile(rf"^第([{NUM}]+)款(?:\s*(.*))?$")
ITEM = re.compile(rf"^[（(]([{NUM}]+)[）)]\s*(.*)$")
ABS_REF = re.compile(rf"第([{NUM}]+)条(?:第([{NUM}]+)款)?(?:第([{NUM}]+)项)?")
NAMED_REF = re.compile(rf"《([^》\n]{{2,100}})》\s*第([{NUM}]+)条(?:第([{NUM}]+)款)?(?:第([{NUM}]+)项)?")
RELATIVE = re.compile(r"本条|本款|前款|上一条|前条")


class PipelineError(RuntimeError):
    pass


def sha256(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(block)
    return value.hexdigest()


def canonical(value: str) -> str:
    return re.sub(r"[\s《》【】()（）\[\]〔〕,，.。:：;；\-—_]", "", value or "").lower()


def chinese_number(value: str) -> str:
    """Keep the displayed legal number stable while normalizing Arabic variants."""
    return re.sub(r"\s+", "", value).replace("〇", "零")


def reliable(text: str) -> bool:
    return len(re.sub(r"\s+", "", text or "")) >= 12 and "�" not in (text or "")


def safe_workspace(value: str) -> Path:
    path = Path(value).expanduser().resolve()
    if path.exists() and any(path.iterdir()):
        raise PipelineError("工作目录必须不存在或为空")
    path.mkdir(parents=True, mode=0o700)
    os.chmod(path, 0o700)
    (path / MARKER).write_text("temporary\n", encoding="utf-8")
    return path


def cleanup(path: Path) -> None:
    path = path.resolve()
    if not (path / MARKER).is_file() or path in {Path("/").resolve(), Path.home().resolve()} or len(path.parts) < 4:
        raise PipelineError("拒绝清理未标记或过于宽泛的目录")
    shutil.rmtree(path)


def managed_python() -> Path:
    root = Path(os.environ.get("CODEX_SKILL_RUNTIME_DIR", Path.home() / ".cache" / "codex-skills" / "legal-structure")).expanduser()
    return root / "venv" / ("Scripts/python.exe" if os.name == "nt" else "bin/python")


def ensure_runtime() -> None:
    """Provide the report writer in the same local managed runtime as MarkItDown."""
    if importlib.util.find_spec("openpyxl"):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        globals().update(Workbook=Workbook, Alignment=Alignment, Font=Font, PatternFill=PatternFill, get_column_letter=get_column_letter)
        return
    python = managed_python()
    python.parent.parent.mkdir(parents=True, exist_ok=True)
    if not python.exists():
        result = subprocess.run([sys.executable, "-m", "venv", str(python.parent.parent)], text=True, capture_output=True, timeout=180)
        if result.returncode:
            raise PipelineError("无法创建报告依赖的受控虚拟环境")
    result = subprocess.run([str(python), "-m", "pip", "install", "--disable-pip-version-check", "openpyxl"], text=True, capture_output=True, timeout=300)
    if result.returncode:
        raise PipelineError("自动安装XLSX报告依赖失败；请检查本机网络或配置离线依赖包后重试")
    os.execv(str(python), [str(python), str(Path(__file__).resolve()), *sys.argv[1:]])


def ensure_markitdown() -> tuple[str, str]:
    python = managed_python()
    candidate = str(python) if python.exists() else sys.executable
    probe = subprocess.run([candidate, "-c", "import markitdown"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    if probe.returncode == 0:
        return candidate, "already_available"
    python.parent.parent.mkdir(parents=True, exist_ok=True)
    if not python.exists():
        result = subprocess.run([sys.executable, "-m", "venv", str(python.parent.parent)], text=True, capture_output=True, timeout=180)
        if result.returncode:
            raise PipelineError("无法创建MarkItDown受控虚拟环境")
    result = subprocess.run([str(python), "-m", "pip", "install", "--disable-pip-version-check", "markitdown"], text=True, capture_output=True, timeout=300)
    if result.returncode:
        raise PipelineError("自动安装MarkItDown失败；请检查本机网络或配置离线依赖包后重试")
    return str(python), "installed_managed_venv"


def safe_extract(archive: Path, destination: Path) -> list[Path]:
    files, total = [], 0
    with zipfile.ZipFile(archive) as container:
        entries = container.infolist()
        if len(entries) > MAX_FILES:
            raise PipelineError("ZIP文件数量超过安全上限")
        for info in entries:
            relative = Path(info.filename)
            if info.is_dir():
                continue
            if relative.is_absolute() or ".." in relative.parts:
                raise PipelineError("ZIP包含路径穿越条目")
            suffix = relative.suffix.lower()
            if suffix in ARCHIVES:
                raise PipelineError("不支持嵌套压缩包")
            if suffix in FORBIDDEN:
                raise PipelineError("ZIP包含可执行或脚本文件")
            if info.file_size > MAX_FILE or (info.compress_size and info.file_size / info.compress_size > MAX_RATIO):
                raise PipelineError("ZIP包含超大文件或异常压缩比")
            total += info.file_size
            if total > MAX_TOTAL:
                raise PipelineError("ZIP解压总量超过安全上限")
            target = (destination / relative).resolve()
            if destination.resolve() not in target.parents:
                raise PipelineError("ZIP目标路径越界")
            target.parent.mkdir(parents=True, exist_ok=True)
            with container.open(info) as source, target.open("wb") as output:
                shutil.copyfileobj(source, output)
            files.append(target)
    return files


def collect(inputs: list[str], workspace: Path) -> list[Path]:
    files: list[Path] = []
    unpacked = workspace / "unpacked"
    unpacked.mkdir()
    for raw in inputs:
        path = Path(raw).expanduser().resolve()
        if not path.exists():
            raise PipelineError(f"输入不存在：{path}")
        candidates = sorted(item for item in path.rglob("*") if item.is_file()) if path.is_dir() else [path]
        for item in candidates:
            suffix = item.suffix.lower()
            if suffix in ARCHIVES:
                files.extend(safe_extract(item, unpacked / f"{len(files):05d}"))
            elif suffix in SUPPORTED:
                files.append(item)
            elif suffix in FORBIDDEN:
                raise PipelineError(f"拒绝可执行或脚本文件：{item.name}")
    if not files:
        raise PipelineError("没有发现支持的法律材料")
    if len(files) > MAX_FILES:
        raise PipelineError("材料数量超过安全上限")
    return files


def markdown_convert(python: str, path: Path) -> tuple[str, str]:
    if path.suffix.lower() in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="replace"), "plain_text"
    code = "from markitdown import MarkItDown; import sys; r=MarkItDown().convert_local(sys.argv[1]); print(getattr(r,'markdown','') or getattr(r,'text_content','') or '')"
    result = subprocess.run([python, "-c", code, str(path)], text=True, capture_output=True, timeout=300)
    return (result.stdout, "MarkItDown convert_local") if result.returncode == 0 else ("", "MarkItDown转换失败")


def ocr(path: Path) -> tuple[str, str]:
    try:
        result = subprocess.run(["tesseract", str(path), "stdout", "-l", "chi_sim+eng"], text=True, capture_output=True, timeout=300)
        if result.returncode == 0 and reliable(result.stdout):
            return result.stdout, "本地Tesseract OCR（待复核）"
    except (FileNotFoundError, subprocess.SubprocessError):
        pass
    return "", "本地OCR不可用或无可靠结果"


def convert(python: str, path: Path) -> tuple[str, str]:
    text, method = markdown_convert(python, path)
    if not reliable(text) and path.suffix.lower() in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".pdf"}:
        text, method = ocr(path)
    return (text, method) if reliable(text) else ("", method or "无法取得可靠文本")


def chunks(text: str, source_id: str) -> list[dict[str, Any]]:
    starts = [match.start() for match in re.finditer(rf"(?m)^第[{NUM}]+(?:编|章|节|条)", text)]
    if not starts:
        starts = list(range(0, len(text), 6000))
    return [{"chunk_id": f"{source_id}-C{index + 1:04d}", "start": start, "end": starts[index + 1] if index + 1 < len(starts) else len(text), "locator": f"提取文本第{text.count(chr(10), 0, start) + 1}行"} for index, start in enumerate(starts)]


def load_json(path: str | Path, name: str) -> dict[str, Any]:
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception as exc:
        raise PipelineError(f"{name}不是有效JSON：{exc}") from exc
    if not isinstance(data, dict):
        raise PipelineError(f"{name}根节点必须是对象")
    return data


def validate_query(query: dict[str, Any]) -> None:
    if not str(query.get("title", "")).strip() or query.get("mode") not in {"structure", "network"}:
        raise PipelineError("query-json必须包含非空title和mode（structure或network）")
    labels = query.get("materials")
    if not isinstance(labels, list) or not labels or any(not isinstance(item, dict) or not str(item.get("file", "")).strip() or not str(item.get("label", "")).strip() for item in labels):
        raise PipelineError("query-json必须包含materials数组；每项需提供file和label")


def parse_title(text: str, fallback: str) -> str:
    for line in text.splitlines()[:30]:
        value = re.sub(r"^#{1,6}\s*", "", line).strip()
        if 3 <= len(value) <= 100 and not re.match(r"^(目录|目\s*录|第[一二三四五六七八九十百千万零〇两0-9]+[编章节条])", value):
            return value
    return fallback


def source_locator(text: str, offset: int) -> str:
    return f"提取文本第{text.count(chr(10), 0, offset) + 1}行"


def node_id(source_id: str, kind: str, number: str, count: int) -> str:
    return f"{source_id}:{kind}-{canonical(number) or 'unnumbered'}-{count:04d}"


def parse_nodes(source: dict[str, Any], text: str) -> list[dict[str, Any]]:
    nodes: list[dict[str, Any]] = []
    stack: dict[str, str] = {}
    order = {"document": 0, "编": 1, "章": 2, "节": 3, "条": 4, "款": 5, "项": 6}
    counters: Counter[str] = Counter()
    source_id = source["source_id"]
    root = {"node_id": f"{source_id}:ROOT", "source_id": source_id, "node_type": "document", "number": "", "title": source["title"], "text": "", "parent_id": "", "path": source["title"], "source_locator": "文件根节点", "order": 0, "confidence": 0.7 if source["ocr_review_required"] else 1.0, "review_status": "待复核" if source["ocr_review_required"] else "已解析"}
    nodes.append(root)
    stack["document"] = root["node_id"]
    current_article = ""
    paragraph_number = 0
    offset = 0
    for raw in text.splitlines(keepends=True):
        line = raw.strip().lstrip("#").strip()
        line_offset = offset
        offset += len(raw)
        if not line or line.startswith("-") and "SHA-256" in line:
            continue
        matched: tuple[str, str, str] | None = None
        for kind, pattern in HEADING:
            found = pattern.match(line)
            if found:
                matched = (kind, chinese_number(found.group(1)), (found.group(2) or "").strip())
                break
        clause = EXPLICIT_CLAUSE.match(line)
        item = ITEM.match(line)
        if clause and current_article:
            matched = ("款", chinese_number(clause.group(1)), (clause.group(2) or "").strip())
        elif item and current_article:
            matched = ("项", chinese_number(item.group(1)), (item.group(2) or "").strip())
        if matched:
            kind, number, body = matched
            counters[kind] += 1
            if kind == "项":
                parent = stack.get("款")
                if not parent:
                    paragraph_number += 1
                    p_id = node_id(source_id, "款", str(paragraph_number), counters["款"] + 1)
                    parent_article = stack.get("条", stack["document"])
                    p_path = next(node["path"] for node in nodes if node["node_id"] == parent_article) + f" / 第{paragraph_number}款"
                    nodes.append({"node_id": p_id, "source_id": source_id, "node_type": "款", "number": str(paragraph_number), "title": "", "text": "", "parent_id": parent_article, "path": p_path, "source_locator": source_locator(text, line_offset), "order": len(nodes), "confidence": root["confidence"], "review_status": root["review_status"]})
                    stack["款"] = p_id
                    parent = p_id
            else:
                parent = stack.get({"编": "document", "章": "编", "节": "章", "条": "节", "款": "条"}[kind])
                if not parent and kind == "条":
                    parent = stack.get("章") or stack.get("编") or stack["document"]
                if not parent and kind == "节":
                    parent = stack.get("章") or stack.get("编") or stack["document"]
                parent = parent or stack["document"]
            parent_node = next(node for node in nodes if node["node_id"] == parent)
            label = f"第{number}{kind}" if kind in {"编", "章", "节", "条", "款"} else f"（{number}）"
            identifier = node_id(source_id, kind, number, counters[kind])
            nodes.append({"node_id": identifier, "source_id": source_id, "node_type": kind, "number": number, "title": body if kind in {"编", "章", "节", "条"} else "", "text": body if kind in {"款", "项"} else "", "parent_id": parent, "path": parent_node["path"] + " / " + label, "source_locator": source_locator(text, line_offset), "order": len(nodes), "confidence": root["confidence"], "review_status": root["review_status"]})
            stack[kind] = identifier
            for lower, level in order.items():
                if order.get(kind, 99) < level:
                    stack.pop(lower, None)
            if kind == "条":
                current_article = identifier
                paragraph_number = 0
                stack.pop("款", None)
                stack.pop("项", None)
            continue
        if current_article:
            if not stack.get("款"):
                paragraph_number += 1
                counters["款"] += 1
                article_node = next(node for node in nodes if node["node_id"] == current_article)
                identifier = node_id(source_id, "款", str(paragraph_number), counters["款"])
                nodes.append({"node_id": identifier, "source_id": source_id, "node_type": "款", "number": str(paragraph_number), "title": "", "text": line, "parent_id": current_article, "path": article_node["path"] + f" / 第{paragraph_number}款", "source_locator": source_locator(text, line_offset), "order": len(nodes), "confidence": root["confidence"], "review_status": root["review_status"]})
                stack["款"] = identifier
            else:
                nodes[-1]["text"] = (nodes[-1]["text"] + "\n" + line).strip()
    return nodes


def resolve_references(nodes: list[dict[str, Any]], sources: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_source: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_id = {node["node_id"]: node for node in nodes}
    for node in nodes:
        by_source[node["source_id"]].append(node)
    article_index: dict[tuple[str, str], list[str]] = defaultdict(list)
    title_index: dict[str, list[str]] = defaultdict(list)
    for source in sources:
        title_index[canonical(source["title"])].append(source["source_id"])
    for node in nodes:
        if node["node_type"] == "条":
            article_index[(node["source_id"], canonical(node["number"]))].append(node["node_id"])
    edges: list[dict[str, Any]] = []
    unresolved: list[dict[str, Any]] = []
    for node in nodes:
        if node["node_type"] == "document":
            continue
        if node["parent_id"]:
            edges.append({"edge_id": f"H-{len(edges)+1:06d}", "edge_type": "层级", "source_node_id": node["parent_id"], "target_node_id": node["node_id"], "reference_text": "", "source_locator": node["source_locator"], "confidence": 1.0, "review_status": node["review_status"]})
        content = " ".join(filter(None, [node["title"], node["text"]]))
        if not content:
            continue
        spans: list[tuple[int, int]] = []
        for match in NAMED_REF.finditer(content):
            spans.append(match.span())
            candidates = title_index.get(canonical(match.group(1)), [])
            article = canonical(match.group(2))
            targets = article_index.get((candidates[0], article), []) if len(candidates) == 1 else []
            _citation_result(edges, unresolved, node, match.group(0), targets, "跨文件绝对引用", "未提供或存在多个同名法规" if len(candidates) != 1 else "未找到唯一条文")
        for match in ABS_REF.finditer(content):
            if any(match.start() >= start and match.end() <= end for start, end in spans):
                continue
            targets = article_index.get((node["source_id"], canonical(match.group(1))), [])
            _citation_result(edges, unresolved, node, match.group(0), targets, "本文件绝对引用", "未找到唯一条文")
        for match in RELATIVE.finditer(content):
            target = relative_target(node, by_id, by_source[node["source_id"]], match.group(0))
            _citation_result(edges, unresolved, node, match.group(0), [target] if target else [], "相对引用", "无法唯一解析相对引用")
    return edges, unresolved


def _citation_result(edges: list[dict[str, Any]], unresolved: list[dict[str, Any]], node: dict[str, Any], text: str, targets: list[str], reference_type: str, reason: str) -> None:
    if len(targets) == 1:
        edges.append({"edge_id": f"C-{sum(edge['edge_type'] == '引用' for edge in edges)+1:06d}", "edge_type": "引用", "source_node_id": node["node_id"], "target_node_id": targets[0], "reference_text": text, "source_locator": node["source_locator"], "confidence": node["confidence"], "review_status": node["review_status"]})
    else:
        unresolved.append({"reference_id": f"U-{len(unresolved)+1:06d}", "source_node_id": node["node_id"], "source_id": node["source_id"], "reference_text": text, "reference_type": reference_type, "reason": reason, "source_locator": node["source_locator"], "confidence": node["confidence"], "review_status": "待复核"})


def relative_target(node: dict[str, Any], by_id: dict[str, dict[str, Any]], local_nodes: list[dict[str, Any]], reference: str) -> str:
    ancestor = node
    if reference == "本条":
        while ancestor and ancestor["node_type"] != "条":
            ancestor = by_id.get(ancestor["parent_id"])
        return ancestor["node_id"] if ancestor else ""
    if reference == "本款":
        while ancestor and ancestor["node_type"] != "款":
            ancestor = by_id.get(ancestor["parent_id"])
        return ancestor["node_id"] if ancestor else ""
    desired = "款" if reference == "前款" else "条"
    while ancestor and ancestor["node_type"] != desired:
        ancestor = by_id.get(ancestor["parent_id"])
    if not ancestor:
        return ""
    siblings = [candidate for candidate in local_nodes if candidate["parent_id"] == ancestor["parent_id"] and candidate["node_type"] == desired and candidate["order"] < ancestor["order"]]
    return max(siblings, key=lambda candidate: candidate["order"])["node_id"] if siblings else ""


def prepare(args: argparse.Namespace) -> None:
    if args.processing_environment == "cloud" and not args.privacy_confirmed:
        raise PipelineError("云端环境必须先取得用户明确知情确认")
    query = load_json(args.query_json, "query-json")
    validate_query(query)
    declarations = {str(item["file"]): item for item in query["materials"]}
    workspace = safe_workspace(args.workspace)
    try:
        markdown_dir, chunk_dir = workspace / "markdown", workspace / "chunks"
        markdown_dir.mkdir(); chunk_dir.mkdir()
        files = collect(args.input, workspace)
        filenames = [path.name for path in files]
        missing = [name for name in filenames if name not in declarations]
        if missing:
            raise PipelineError("query-json的materials缺少文件标签：" + "、".join(missing))
        if any(path.suffix.lower() not in {".txt", ".md"} for path in files):
            python, dependency_status = ensure_markitdown()
        else:
            python, dependency_status = sys.executable, "not_required_plain_text"
        sources, seen = [], {}
        for index, path in enumerate(files, 1):
            source_id = f"SRC-{index:05d}"
            digest = sha256(path)
            declaration = declarations[path.name]
            text, conversion = convert(python, path)
            duplicate = seen.get(digest)
            seen.setdefault(digest, source_id)
            status = "重复" if duplicate else ("成功" if text else "无法读取")
            title = str(declaration.get("title") or parse_title(text, path.stem))
            record = {"source_id": source_id, "source_name": path.name, "source_path": str(path), "source_sha256": digest, "label": str(declaration["label"]), "version_label": str(declaration.get("version_label", "")), "title": title, "format": path.suffix.lower(), "status": status, "duplicate_of": duplicate or "", "conversion": conversion, "char_count": len(text), "ocr_review_required": "OCR" in conversion, "markdown_path": "", "chunks_path": ""}
            if text and not duplicate:
                markdown = markdown_dir / f"{source_id}.md"
                markdown.write_text(f"# {title}\n\n- 材料ID：{source_id}\n- 用户标签：{record['label']}\n- SHA-256：{digest}\n- 转换：{conversion}\n\n{text}", encoding="utf-8")
                os.chmod(markdown, 0o600)
                chunk_file = chunk_dir / f"{source_id}.json"
                chunk_file.write_text(json.dumps(chunks(text, source_id), ensure_ascii=False, indent=2), encoding="utf-8")
                os.chmod(chunk_file, 0o600)
                record.update(markdown_path=str(markdown), chunks_path=str(chunk_file))
            sources.append(record)
        index = {"schema": SCHEMA, "query": query, "dependency_status": dependency_status, "sources": sources, "reading_policy": "先读索引，再按文件与条文切片读取；不得默认加载全部Markdown"}
        index_path = workspace / "legal-structure-index.json"
        index_path.write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")
        bundle = {"schema": SCHEMA, "created_at": datetime.now().astimezone().isoformat(timespec="seconds"), "query": query, "dependency_status": dependency_status, "index_path": str(index_path), "sources": sources}
        output = workspace / "legal-structure-bundle.json"
        output.write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")
        print(output)
    except Exception:
        cleanup(workspace)
        raise


def style(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for index in range(1, sheet.max_column + 1):
        width = max(len(str(sheet.cell(row, index).value or "")) for row in range(1, min(sheet.max_row, 100) + 1)) + 2
        sheet.column_dimensions[get_column_letter(index)].width = min(48, max(12, width))


def add_sheet(book: Workbook, name: str, columns: list[str], values: list[dict[str, Any]]) -> None:
    sheet = book.create_sheet(name)
    sheet.append(columns)
    for item in values:
        sheet.append([json.dumps(item.get(column), ensure_ascii=False) if isinstance(item.get(column), (list, dict)) else item.get(column, "") for column in columns])
    style(sheet)


def validate_analysis(data: dict[str, Any], source_ids: set[str], node_ids: set[str]) -> tuple[dict[str, str], list[dict[str, str]]]:
    overrides: dict[str, str] = {}
    for item in data.get("source_identity_overrides", []):
        if not isinstance(item, dict) or item.get("source_id") not in source_ids or not str(item.get("title", "")).strip() or not str(item.get("reason", "")).strip():
            raise PipelineError("source_identity_overrides包含无效记录")
        overrides[item["source_id"]] = str(item["title"])
    review: list[dict[str, str]] = []
    for item in data.get("review_items", []):
        if not isinstance(item, dict) or item.get("source_id") not in source_ids or item.get("node_id") not in node_ids or not str(item.get("reason", "")).strip():
            raise PipelineError("review_items包含未知来源、节点或缺少原因")
        review.append({"item_id": f"A-{len(review)+1:05d}", "source_id": str(item["source_id"]), "node_id": str(item["node_id"]), "field": str(item.get("field", "")), "reason": str(item["reason"]), "source_locator": "用户复核"})
    return overrides, review


def mermaid_label(node: dict[str, Any]) -> str:
    text = (node["path"].split(" / ")[-1] + (" " + node["title"] if node["title"] else "")).replace('"', "'")
    return text[:80]


def mermaid_id(value: str) -> str:
    return "N" + hashlib.sha1(value.encode("utf-8")).hexdigest()[:12]


def write_mermaid(directory: Path, sources: list[dict[str, Any]], nodes: list[dict[str, Any]], edges: list[dict[str, Any]]) -> list[dict[str, str]]:
    directory.mkdir(parents=True)
    manifest = []
    for source in sources:
        selected = [node for node in nodes if node["source_id"] == source["source_id"]]
        if not selected:
            continue
        groups = [selected[index:index + 60] for index in range(0, len(selected), 60)]
        for part, group in enumerate(groups, 1):
            ids = {node["node_id"] for node in group}
            lines = ["```mermaid", "flowchart TD"]
            for node in group:
                lines.append(f'  {mermaid_id(node["node_id"])}["{mermaid_label(node)}"]')
            for edge in edges:
                if edge["edge_type"] == "层级" and edge["source_node_id"] in ids and edge["target_node_id"] in ids:
                    lines.append(f'  {mermaid_id(edge["source_node_id"])} --> {mermaid_id(edge["target_node_id"])}')
            lines.append("```")
            filename = f"{source['source_id']}_{part:02d}.mmd"
            (directory / filename).write_text("\n".join(lines), encoding="utf-8")
            manifest.append({"source_id": source["source_id"], "source_title": source["title"], "file": filename, "node_count": str(len(group)), "scope": f"第{part}段（每段最多60节点）"})
    return manifest


def html_report(path: Path, title: str, mode: str, nodes: list[dict[str, Any]], edges: list[dict[str, Any]], unresolved: list[dict[str, Any]]) -> None:
    payload = json.dumps({"nodes": nodes, "edges": edges, "unresolved": unresolved, "mode": mode}, ensure_ascii=False).replace("</", "<\\/")
    escaped_title = html.escape(title)
    page = f"""<!doctype html><html lang=\"zh-CN\"><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1\"><title>{escaped_title} 法律条文结构图谱</title>
<style>body{{margin:0;font:14px -apple-system,BlinkMacSystemFont,'PingFang SC',sans-serif;color:#18212b;background:#f6f8fa}}header{{background:#103d5c;color:#fff;padding:18px 24px}}h1{{font-size:20px;margin:0 0 6px}}main{{display:grid;grid-template-columns:minmax(320px,42%) 1fr;gap:14px;padding:14px}}section{{background:#fff;border:1px solid #d8e0e6;border-radius:6px;padding:14px;min-width:0}}input,select{{box-sizing:border-box;width:100%;padding:8px;border:1px solid #9eabb5;border-radius:4px;margin:0 0 10px}}ul{{list-style:none;margin:0;padding-left:17px}}li{{margin:4px 0}}summary{{cursor:pointer}}button{{border:0;background:transparent;color:#0b5e91;text-align:left;padding:2px;cursor:pointer}}button:hover{{text-decoration:underline}}.meta{{color:#667786;font-size:12px}}.edge{{border-bottom:1px solid #e5eaee;padding:9px 0}}.unresolved{{color:#9c3b19}}#detail{{white-space:pre-wrap;line-height:1.55}}#graph{{width:100%;height:280px;border:1px solid #d8e0e6;background:#fbfcfd}}.edge-line{{stroke:#9fb8c7;stroke-width:1.5}}.graph-node{{fill:#e1f0f7;stroke:#17628c;stroke-width:1.2}}.graph-label{{font-size:10px;fill:#18212b}}@media(max-width:800px){{main{{grid-template-columns:1fr}}}}</style>
<header><h1>{escaped_title}法律条文结构图谱</h1><div>模式：{'结构导图' if mode == 'structure' else '引用网络'}；仅分析用户提交材料。</div></header><main><section><input id=\"search\" placeholder=\"搜索条号、标题或关键词\"><div id=\"tree\"></div></section><section><select id=\"kind\"><option value=\"all\">全部关系</option><option value=\"引用\">仅条文引用</option><option value=\"层级\">仅层级关系</option></select><svg id=\"graph\" viewBox=\"0 0 800 280\" aria-label=\"条文关系图\"></svg><div id=\"detail\">选择左侧节点查看原文与来源定位。</div><h3>关系</h3><div id=\"edges\"></div><h3>未解析引用</h3><div id=\"unresolved\"></div></section></main>
<script>const data={payload};const byId=Object.fromEntries(data.nodes.map(n=>[n.node_id,n]));const children={{}};data.nodes.forEach(n=>{{(children[n.parent_id]??=[]).push(n)}});Object.values(children).forEach(a=>a.sort((x,y)=>x.order-y.order));const label=n=>`${{n.path.split(' / ').at(-1)}}${{n.title?' '+n.title:''}}`;
function renderTree(){{const q=document.querySelector('#search').value.trim().toLowerCase();function walk(id){{const items=(children[id]||[]).filter(n=>!q||JSON.stringify(n).toLowerCase().includes(q));if(!items.length)return '';return '<ul>'+items.map(n=>{{const sub=walk(n.node_id);return `<li>${{sub?'<details open><summary>':''}}<button data-id="${{n.node_id}}">${{label(n)}}</button>${{sub?'</summary>'+sub+'</details>':''}}</li>`}}).join('')+'</ul>'}}document.querySelector('#tree').innerHTML=walk('')}}
function show(id){{const n=byId[id];document.querySelector('#detail').textContent=`${{label(n)}}\n\n来源：${{n.source_locator}}\n路径：${{n.path}}\n\n文本：${{n.text||'（结构节点无独立正文）'}}`;renderEdges(id)}}
function renderEdges(id=''){{const kind=document.querySelector('#kind').value;const list=data.edges.filter(e=>(kind==='all'||e.edge_type===kind)&&(!id||e.source_node_id===id||e.target_node_id===id));document.querySelector('#edges').innerHTML=list.map(e=>`<div class=edge><b>${{e.edge_type}}</b>：${{label(byId[e.source_node_id])}} → ${{label(byId[e.target_node_id])}}<div class=meta>${{e.reference_text||e.source_locator}}</div></div>`).join('')||'<div class=meta>没有符合筛选条件的关系。</div>';document.querySelector('#unresolved').innerHTML=data.unresolved.map(u=>`<div class=unresolved>${{u.reference_text}}：${{u.reason}}<div class=meta>${{u.source_locator}}</div></div>`).join('')||'<div class=meta>无。</div>';renderGraph(list)}}
function renderGraph(edges){{const svg=document.querySelector('#graph');svg.replaceChildren();const chosen=edges.slice(0,80), ids=[...new Set(chosen.flatMap(e=>[e.source_node_id,e.target_node_id]))].slice(0,50), position={{}};ids.forEach((id,i)=>{{position[id]={{x:70+(i%5)*165,y:35+Math.floor(i/5)*48}}}});const ns=['http:','','www.w3.org','2000','svg'].join('/');for(const e of chosen){{if(!position[e.source_node_id]||!position[e.target_node_id])continue;const a=position[e.source_node_id],b=position[e.target_node_id],line=document.createElementNS(ns,'line');line.setAttribute('x1',a.x);line.setAttribute('y1',a.y);line.setAttribute('x2',b.x);line.setAttribute('y2',b.y);line.setAttribute('class','edge-line');svg.append(line)}}ids.forEach(id=>{{const p=position[id],g=document.createElementNS(ns,'g'),circle=document.createElementNS(ns,'circle'),text=document.createElementNS(ns,'text');circle.setAttribute('cx',p.x);circle.setAttribute('cy',p.y);circle.setAttribute('r','12');circle.setAttribute('class','graph-node');text.setAttribute('x',p.x+15);text.setAttribute('y',p.y+4);text.setAttribute('class','graph-label');text.textContent=label(byId[id]).slice(0,22);g.append(circle,text);g.addEventListener('click',()=>show(id));svg.append(g)}})}}
document.querySelector('#tree').addEventListener('click',e=>{{if(e.target.dataset.id)show(e.target.dataset.id)}});document.querySelector('#search').addEventListener('input',renderTree);document.querySelector('#kind').addEventListener('change',()=>renderEdges());renderTree();renderEdges();</script></html>"""
    path.write_text(page, encoding="utf-8")


def report(args: argparse.Namespace) -> None:
    bundle_path = Path(args.bundle).expanduser().resolve()
    workspace = bundle_path.parent
    if not (workspace / MARKER).is_file():
        raise PipelineError("bundle不在受管工作目录")
    bundle = load_json(bundle_path, "bundle")
    sources = [source for source in bundle["sources"] if source["status"] == "成功"]
    nodes: list[dict[str, Any]] = []
    for source in sources:
        markdown = Path(source["markdown_path"])
        text = markdown.read_text(encoding="utf-8", errors="replace")
        body = text.split("\n\n", 4)[-1] if "材料ID" in text else text
        nodes.extend(parse_nodes(source, body))
    source_ids, node_ids = {source["source_id"] for source in sources}, {node["node_id"] for node in nodes}
    analysis = load_json(args.analysis_json, "analysis-json")
    overrides, review_items = validate_analysis(analysis, source_ids, node_ids)
    for source in sources:
        if source["source_id"] in overrides:
            source["title"] = overrides[source["source_id"]]
    for node in nodes:
        if node["node_type"] == "document":
            node["title"] = next(source["title"] for source in sources if source["source_id"] == node["source_id"])
    edges, unresolved = resolve_references(nodes, sources)
    for item in review_items:
        target = next(node for node in nodes if node["node_id"] == item["node_id"])
        target["review_status"] = "待复核"
    for source in bundle["sources"]:
        if source["status"] != "成功":
            review_items.append({"item_id": f"A-{len(review_items)+1:05d}", "source_id": source["source_id"], "node_id": "", "field": "材料读取", "reason": source["conversion"] if source["status"] == "无法读取" else f"与{source['duplicate_of']}内容重复", "source_locator": source["source_name"]})
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    base = re.sub(r'[\\/:*?"<>|]', "_", str(bundle["query"]["title"]))
    html_path, xlsx_path, mermaid_dir = output_dir / f"{base}_法律条文结构图谱.html", output_dir / f"{base}_法律条文结构图谱.xlsx", output_dir / f"{base}_Mermaid"
    if html_path.exists() or xlsx_path.exists() or mermaid_dir.exists():
        raise PipelineError("输出已存在，拒绝覆盖")
    mermaid = write_mermaid(mermaid_dir, sources, nodes, edges)
    html_report(html_path, str(bundle["query"]["title"]), str(bundle["query"]["mode"]), nodes, edges, unresolved)
    book = Workbook(); overview = book.active; overview.title = "图谱概览"; overview.append(["项目", "内容"])
    overview_rows = [("主题", bundle["query"]["title"]), ("模式", bundle["query"]["mode"]), ("成功材料数", len(sources)), ("条文节点数", len(nodes)), ("层级关系数", sum(edge["edge_type"] == "层级" for edge in edges)), ("条文引用数", sum(edge["edge_type"] == "引用" for edge in edges)), ("未解析引用数", len(unresolved)), ("说明", "仅分析用户提交材料；未解析项不构成确定关系。")]
    for row in overview_rows: overview.append(row)
    style(overview)
    processing = [{"source_id": source["source_id"], "source_file": source["source_name"], "label": source["label"], "version_label": source["version_label"], "title": source["title"], "sha256": source["source_sha256"], "format": source["format"], "status": source["status"], "conversion": source["conversion"], "char_count": source["char_count"], "ocr_review_required": source["ocr_review_required"], "duplicate_of": source["duplicate_of"], "dependency_status": bundle["dependency_status"]} for source in bundle["sources"]]
    add_sheet(book, "材料与处理记录", list(processing[0].keys()) if processing else ["source_id"], processing)
    add_sheet(book, "条文节点", ["node_id", "source_id", "node_type", "number", "title", "text", "parent_id", "path", "source_locator", "order", "confidence", "review_status"], nodes)
    add_sheet(book, "层级关系", ["edge_id", "source_node_id", "target_node_id", "source_locator", "confidence", "review_status"], [{key: value for key, value in edge.items() if key != "reference_text"} for edge in edges if edge["edge_type"] == "层级"])
    add_sheet(book, "条文引用", ["edge_id", "source_node_id", "target_node_id", "reference_text", "source_locator", "confidence", "review_status"], [{key: value for key, value in edge.items() if key != "edge_type"} for edge in edges if edge["edge_type"] == "引用"])
    add_sheet(book, "未解析引用", ["reference_id", "source_node_id", "source_id", "reference_text", "reference_type", "reason", "source_locator", "confidence", "review_status"], unresolved)
    add_sheet(book, "待复核", ["item_id", "source_id", "node_id", "field", "reason", "source_locator"], review_items)
    summary = [{"metric": "节点类型：" + kind, "value": count} for kind, count in sorted(Counter(node["node_type"] for node in nodes).items())] + [{"metric": "待复核节点", "value": sum(node["review_status"] == "待复核" for node in nodes)}, {"metric": "未解析引用", "value": len(unresolved)}]
    add_sheet(book, "统计汇总", ["metric", "value"], summary)
    add_sheet(book, "Mermaid清单", ["source_id", "source_title", "file", "node_count", "scope"], mermaid)
    book.save(xlsx_path); book.close()
    if args.cleanup:
        cleanup(workspace)
    print(json.dumps({"html": str(html_path), "xlsx": str(xlsx_path), "mermaid_dir": str(mermaid_dir)}, ensure_ascii=False))


def main() -> int:
    ensure_runtime()
    parser = argparse.ArgumentParser()
    commands = parser.add_subparsers(dest="command", required=True)
    prepare_parser = commands.add_parser("prepare")
    prepare_parser.add_argument("--input", nargs="+", required=True); prepare_parser.add_argument("--query-json", required=True); prepare_parser.add_argument("--workspace", required=True); prepare_parser.add_argument("--processing-environment", choices=["local", "cloud"], required=True); prepare_parser.add_argument("--privacy-confirmed", action="store_true"); prepare_parser.set_defaults(func=prepare)
    report_parser = commands.add_parser("report")
    report_parser.add_argument("--bundle", required=True); report_parser.add_argument("--analysis-json", required=True); report_parser.add_argument("--output-dir", required=True); report_parser.add_argument("--cleanup", action="store_true"); report_parser.set_defaults(func=report)
    args = parser.parse_args()
    try:
        args.func(args)
        return 0
    except (PipelineError, OSError, ValueError, KeyError, json.JSONDecodeError, zipfile.BadZipFile, subprocess.SubprocessError) as exc:
        if args.command == "report" and args.cleanup:
            try:
                workspace = Path(args.bundle).expanduser().resolve().parent
                if (workspace / MARKER).is_file(): cleanup(workspace)
            except Exception:
                pass
        print(f"错误：{exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
