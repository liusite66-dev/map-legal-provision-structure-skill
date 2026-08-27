import hashlib
import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RUN = ROOT / "scripts" / "run_pipeline.py"
FIX = ROOT / "tests" / "fixtures"


class PipelineTests(unittest.TestCase):
    def command(self, *args: object, ok: bool = True) -> subprocess.CompletedProcess[str]:
        result = subprocess.run([sys.executable, str(RUN), *map(str, args)], text=True, capture_output=True)
        if ok and result.returncode:
            self.fail(result.stderr)
        return result

    def test_end_to_end_tree_network_and_cleanup(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            corpus = root / "corpus"; corpus.mkdir()
            for name in ("法规甲.txt", "法规乙.txt"):
                (corpus / name).write_bytes((FIX / name).read_bytes())
            before = hashlib.sha256((corpus / "法规甲.txt").read_bytes()).hexdigest()
            bundle = Path(self.command("prepare", "--input", corpus, "--query-json", FIX / "query.json", "--workspace", root / "work", "--processing-environment", "local").stdout.strip())
            index = json.loads((root / "work" / "legal-structure-index.json").read_text())
            self.assertEqual(index["dependency_status"], "not_required_plain_text")
            self.assertEqual(len(index["sources"]), 2)
            analysis = root / "analysis.json"; analysis.write_text(json.dumps({"source_identity_overrides": [], "review_items": []}), encoding="utf-8")
            output = json.loads(self.command("report", "--bundle", bundle, "--analysis-json", analysis, "--output-dir", root / "out", "--cleanup").stdout)
            self.assertFalse((root / "work").exists())
            self.assertEqual(before, hashlib.sha256((corpus / "法规甲.txt").read_bytes()).hexdigest())
            self.assertTrue(Path(output["html"]).is_file())
            self.assertTrue(list(Path(output["mermaid_dir"]).glob("*.mmd")))
            page = Path(output["html"]).read_text(encoding="utf-8")
            self.assertIn('<svg id="graph"', page)
            self.assertIn("<details open>", page)
            book = load_workbook(output["xlsx"], data_only=True)
            self.assertIn("条文节点", book.sheetnames)
            self.assertIn("条文引用", book.sheetnames)
            self.assertGreater(book["条文引用"].max_row, 1)
            self.assertGreater(book["未解析引用"].max_row, 1)
            book.close()

    def test_cloud_confirmation_precedes_input_read(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            result = self.command("prepare", "--input", FIX / "法规甲.txt", "--query-json", FIX / "query.json", "--workspace", root / "work", "--processing-environment", "cloud", ok=False)
            self.assertEqual(result.returncode, 2)
            self.assertFalse((root / "work").exists())

    def test_zip_path_traversal_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw); archive = root / "unsafe.zip"
            with zipfile.ZipFile(archive, "w") as container:
                container.writestr("../法规甲.txt", "x")
            result = self.command("prepare", "--input", archive, "--query-json", FIX / "query.json", "--workspace", root / "work", "--processing-environment", "local", ok=False)
            self.assertEqual(result.returncode, 2)
            self.assertIn("路径穿越", result.stderr)

    def test_unknown_review_node_rejected_and_workspace_cleaned(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw); corpus = root / "corpus"; corpus.mkdir()
            for name in ("法规甲.txt", "法规乙.txt"):
                (corpus / name).write_bytes((FIX / name).read_bytes())
            bundle = Path(self.command("prepare", "--input", corpus, "--query-json", FIX / "query.json", "--workspace", root / "work", "--processing-environment", "local").stdout.strip())
            analysis = root / "bad.json"; analysis.write_text(json.dumps({"review_items": [{"source_id": "SRC-00001", "node_id": "missing", "reason": "bad"}]}), encoding="utf-8")
            result = self.command("report", "--bundle", bundle, "--analysis-json", analysis, "--output-dir", root / "out", "--cleanup", ok=False)
            self.assertEqual(result.returncode, 2)
            self.assertFalse((root / "work").exists())


if __name__ == "__main__":
    unittest.main()
